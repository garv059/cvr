import os
import re
import docx2txt
from flask import Flask, render_template, request, redirect, url_for, send_file
from openpyxl import load_workbook, Workbook
from PyPDF2 import PdfReader
import pythoncom
import win32com.client

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'docx', 'pdf', 'doc'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_information(text):
    email = re.findall(r'\b[A-Za-z0-9._%+-]+ ?@[A-Za-z0-9.-]+ ?\.[A-Z|a-z]{2,}\b', text)
    phone = re.findall(r'\b\d{3}[-.\s]??\d{3}[-.\s]??\d{4}\b', text)
    return email[0] if email else None, phone[0] if phone else None

def extract_text_from_pdf(file_path):
    text = ""
    with open(file_path, 'rb') as f:
        reader = PdfReader(f)
        for page_num in range(len(reader.pages)):
            text += reader.pages[page_num].extract_text()
    return text

def extract_text_from_doc(file_path):
    text = ""
    pythoncom.CoInitialize()
    word = win32com.client.Dispatch("Word.Application")
    word.Visible = False
    doc = word.Documents.Open(file_path)
    for paragraph in doc.Paragraphs:
        
        cleaned_text = ''.join(char for char in paragraph.Range.Text if ord(char) < 128)
        
        cleaned_text = ''.join(char if char.isprintable() else '' for char in cleaned_text)
        text += cleaned_text + '\n' 
    doc.Close()
    word.Quit()
    pythoncom.CoUninitialize()
    return text

@app.route('/download_output', methods=['GET'])
def download_output():
    output_filename = os.path.join(app.config['UPLOAD_FOLDER'], 'output.xlsx')
    return send_file(output_filename, as_attachment=True)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/parse_cvs', methods=['POST'])
def parse_cvs():
    output_filename = os.path.join(app.config['UPLOAD_FOLDER'], 'output.xlsx')

    if os.path.exists(output_filename):
        wb = load_workbook(output_filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Email', 'Contact Number', 'Text'])

    for file in request.files.getlist('cv_files'):
        if file and allowed_file(file.filename):
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(file_path)
            if file.filename.endswith('.docx'):
                text = docx2txt.process(file_path)
            elif file.filename.endswith('.pdf'):
                text = extract_text_from_pdf(file_path)
            elif file.filename.endswith('.doc'):
                text = extract_text_from_doc(file_path)
            else:
                continue
            os.remove(file_path)

            email, phone = extract_information(text)
            text_lines = text.split('\n')
            text_concatenated = ' '.join(text_lines)
            ws.append([email, phone, text_concatenated])

    wb.save(output_filename)
    return redirect(url_for('download_output'))

if __name__ == '__main__':
    app.run(debug=True)
